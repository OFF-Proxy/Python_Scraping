from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# 10行×10列にインクリメントした数字を設定

current_number = 1

for row in range(1, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col).value = current_number
        current_number += 1

# Save the file
wb.save("sample.xlsx")

ws.create_sheet()